# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook, Workbook # ���� �ҷ�����
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# �ʿ� ���� Open
CC_424 = load_workbook("424W_CC_2023-01-06.XLSX")
CC_424_ws = CC_424.active
WBS_424 = load_workbook("424W_WBS_2023-01-06.XLSX")
WBS_424_ws = WBS_424.active
Ref = load_workbook("Reference.xlsx")
Ref_WBS = Ref['WBS']
Ref_CC_465_ws = Ref['465W_CC List']
Ref_CC_424_ws = Ref['424W_CC List']


# Merge ���� ����
wb = Workbook()
ws = wb.active
ws.title = 'Expense_Actual_Update'
# Merge ���� Title ����
ws['A1'] = 'Cost Center';ws['B1'] = 'Cost Element';ws['C1'] = 'Val/COArea Crcy';ws['D1'] = 'Partner object';ws['E1'] = 'Name';ws['F1'] = 'Org Cost Center';ws['G1'] = 'Partner object type'

##### 424W WBS �� ��ȯ �� Merge

# WBS 424�� Profit Center �� �ݺ�
for cell in WBS_424_ws["A"]:
    # title ����
    if cell.row == 1:
        continue
    # PC�� 60040�̸�
    if cell.value == '60040':
        # WBS element�� �տ��� 11������ ����
        WBS_element = WBS_424_ws["D"+cell.coordinate[1:]].value[:11]
        # WBS ����ǥ �� �ݺ�
        for wbs in Ref_WBS['A']:
             #title ����
            if wbs.row == 1:
                continue
            # ����ǥ�� wbs code�� update ������ wbs code�� ������
            if wbs.value == WBS_element:
                # ����ǥ�� CC �� ����
                CC = Ref_WBS["B"+wbs.coordinate[1:]].value
                ws["A"+cell.coordinate[1:]] = CC
    # PC�� 60040�� �ƴϸ� �״�� PC ��������
    else:
        ws["A"+cell.coordinate[1:]] = cell.value
    ws["F"+cell.coordinate[1:]] = cell.value

for cols in WBS_424_ws["B:D"]:
    for cell in cols:
        if cell.row == 1:
            continue
        elif cell.column == 2:
            ws["B"+cell.coordinate[1:]] = cell.value
        elif cell.column == 3:
            ws["C"+cell.coordinate[1:]] = cell.value
            ws["C"+cell.coordinate[1:]].number_format = openpyxl.styles.numbers.builtin_format_code(3)  # '#,##0'
        elif cell.column == 4:
            ws["D"+cell.coordinate[1:]] = cell.value

# ������ �� ���� �߰��ؼ� �����͸� �Է��ϱ� ���� �ε��� ����
i = ws.max_row

##### 424W CC �� ��ȯ �� Merge
for cell_CC in CC_424_ws["A"]:
    # title ����
    if cell_CC.row == 1:
        continue
    # WBS ����ǥ �� �ݺ�
    for CC in Ref_CC_424_ws['A']:
        #title ����
        if CC.row == 1:
            continue
        # type ����
        CC.value = str(CC.value)
        # ����ǥ�� PCCC�� update ������ Cost Center�� ������
        if cell_CC.value == CC.value:
            # ����ǥ�� CC �� ����
            CC_final = Ref_CC_424_ws["B"+CC.coordinate[1:]].value
            ws.cell(row = i, column=1).value = CC_final  #A��(Column=1)�� ���� �ٲٸ鼭 �Է�
            i+=1  #�� �� ����

##### �� �ʺ� ����
for column_cells in ws.columns:
    length = max(len(str(cell.value))*1.1 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length

wb.save('Expense_Actual_Update.xlsx')
wb.close()
